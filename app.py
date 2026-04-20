from flask import Flask, render_template, request, redirect, jsonify, send_file
from PIL import Image
from weasyprint import HTML
from PyPDF2 import PdfMerger
from io import BytesIO
from collections import OrderedDict, defaultdict
from threading import Thread
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as ExcelImage
import openpyxl
import psutil
import traceback
import tempfile
import gc
import io
import local
import Email

app = Flask(__name__)
app.secret_key = 'ddd'

app.config['UPLOAD_FOLDER'] = './static/cache'
app.config['MAX_CONTENT_LENGTH'] = 100 * 3000 * 3000  # 100 MB limit
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}

PASSWORD = 'HV3G849'


def print_memory_usage():
    process = psutil.Process()
    memory_in_mb = process.memory_info().rss / 1024 ** 2
    print(f"Current memory usage: {memory_in_mb:.2f} MB")


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


# Service worker setup
@app.after_request
def add_security_headers(response):
    if request.path == '/static/js/service-worker.js':
        response.headers['Service-Worker-Allowed'] = '/'
        response.headers['Cache-Control'] = 'no-store'
    return response


@app.route('/api/init', methods=['POST'])
def verify_password():
    data = request.json
    if data.get('password') == PASSWORD:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False})


@app.route('/verify')
def verify():
    return render_template('vertify.html')


@app.route('/access')
def access():
    return render_template('access.html')


@app.route('/')
def home():
    return render_template('index.html')


@app.route('/customer-Add')
def add_Customer_Page():
    return render_template('customerAdd.html')


@app.route('/customer')
def view_customers():
    return render_template('customer.html')


@app.route('/customer/edit', methods=['GET', 'POST'])
def edit_customer():
    return render_template('customerEdit.html')


@app.route('/inspection-Details', methods=['GET'])
def inspection_details():
    return render_template('inspectionDetails.html')


@app.route('/inspectionDetails-Add', methods=['GET', 'POST'])
def inspection_detail_add():
    inspection_id = request.args.get('inspection_id')
    return render_template('/inspectionDetailAdd.html', inspection_id=inspection_id)


@app.route('/inspection-Details/edit', methods=['GET', 'POST'])
def edit_inspection_detail():
    return render_template('inspectionDetailEdit.html')


@app.route('/inspections')
def inspection_summary():
    return render_template('inspections.html')


@app.route('/inspection-Add')
def Inspection_add():
    customers = local.fetch('Customer', exclude_image_url=True)
    return render_template('/inspectionsAdd.html', customers=customers)


@app.route('/inspections/edit', methods=['GET', 'POST'])
def edit_inspection():
    return render_template('inspectionsEdit.html')


@app.route('/select-Inspections', methods=['GET', 'POST'])
def select_inspections():
    inspections = local.fetch('Inspection_Header')
    customers = local.fetch('Customer')
    customer_map = {c['customer_id']: c['name'] for c in customers}

    all_details = local.fetch('Inspection_Details')
    details_map = defaultdict(list)
    for detail in all_details:
        details_map[detail['inspection_id']].append(detail)

    filtered_inspections = []
    for inspection in inspections:
        inspection_id = inspection.get('inspection_id')
        details_count = len(details_map[inspection_id])
        inspection['details_count'] = details_count
        inspection['customer_name'] = customer_map.get(inspection['customer_id'], 'Unknown')
        if details_count > 0:
            filtered_inspections.append(inspection)

    return render_template('selectPrint.html', inspections=filtered_inspections)


@app.route('/get_revisions/<int:inspection_id>')
def get_revisions(inspection_id):
    revisions = local.fetch('Revisions', {'inspection_id': inspection_id})
    return render_template('partials/revision_table.html', revisions=revisions)


@app.route('/inspection-Print/<int:inspection_id>', methods=['POST'])
def generate_report(inspection_id):
    try:
        date_issued = request.form.get('date_issued')
        version = request.form.get('version')
        issued_by = request.form.get('issued_by')
        other_version = request.form.get('other_version') if version == 'other' else None

        revisions_data = {
            "inspection_id": inspection_id,
            "date": date_issued,
            "status": other_version if other_version else version,
            "detail": other_version if other_version else version,
            "issued_by": issued_by
        }

        add = request.form.get('add')
        if add:
            try:
                local.insert('Revisions', revisions_data)
                print("Data inserted successfully.")
            except Exception as e:
                print("Error during insert:", e)

        # Step 1: Fetch all data
        inspection_header = local.fetch('Inspection_Header', {'inspection_id': inspection_id})
        if not inspection_header:
            return redirect('/select-Inspections')
        inspection_header = inspection_header[0]

        customer_data = local.fetch('Customer', {'customer_id': inspection_header.get('customer_id')})
        if not customer_data:
            return "Customer not found", 404
        customer = customer_data[0]

        risk_type = customer.get('risk_type')
        inspection_details_raw = local.fetch('Inspection_Details', {'inspection_id': inspection_id})
        inspection_details = sorted(calculate_risk(inspection_details_raw, risk_type), key=lambda x: x['detail_id'])

        # Step 2: Prepare report data
        logo = 'https://zmusspsqfcmjpqnwkpmx.supabase.co/storage/v1/object/public/images//hv.png'
        report_data = {
            "customer": customer,
            "inspection_header": inspection_header,
            "rows": inspection_details,
            "revisions": local.fetch('Revisions', {'inspection_id': inspection_id}),
            "logo": logo
        }

        ROWS_PER_TABLE_PAGE = 15
        row_chunks = [inspection_details[i:i+ROWS_PER_TABLE_PAGE] for i in range(0, len(inspection_details), ROWS_PER_TABLE_PAGE)]
        total_pages = 1 + len(row_chunks) + len(inspection_details)

        pages = [render_template('report-title.html', **report_data, num=f'page 1 of {total_pages}')]

        for i, chunk in enumerate(row_chunks, start=2):
            pages.append(render_template('report-table.html', **{**report_data, 'rows': chunk}, num=f'page {i} of {total_pages}'))

        for i, row in enumerate(inspection_details, start=2 + len(row_chunks)):
            pages.append(render_template('report-detail.html', row=row, **report_data, num=f'page {i} of {total_pages}'))

        # Step 3: Generate PDF
        pdf_files = []
        for page_html in pages:
            pdf_bytes = HTML(string=page_html).write_pdf(resolution=72)
            pdf_files.append(BytesIO(pdf_bytes))

        # Step 4: Merge PDFs
        merged_pdf = PdfMerger()
        for pdf in pdf_files:
            pdf.seek(0)
            merged_pdf.append(pdf)

        final_pdf_buffer = BytesIO()
        merged_pdf.write(final_pdf_buffer)
        final_pdf_buffer.seek(0)
        merged_pdf.close()
        print("PDF merged successfully.")

        inspection_date = inspection_header.get('date')

        # Step 5: Generate Excel with matching column order to PDF
        data_to_pass_to_excel = []
        for row in inspection_details:
            new_row = OrderedDict([('date', inspection_date)])
            new_row.update(row)
            data_to_pass_to_excel.append(new_row)

        excel = generate_excel(data_to_pass_to_excel)

        # Step 6: Send email in background
        email = request.form.get('email')
        customer_name = customer.get('name')
        final_pdf_bytes = final_pdf_buffer.getvalue()

        def send_email_task():
            Email.send_Email(final_pdf_bytes, excel, email, customer_name)

        email_thread = Thread(target=send_email_task)
        email_thread.start()

        del excel, customer
        gc.collect()

        return render_template('report_ready.html', email=email)

    except Exception as e:
        traceback.print_exc()
        return redirect('/')


def calculate_risk(data, risk_type):
    """
    Calculates a risk rating for each record based on 'probability' and 'consequence'.
    """
    probability_map = {
        'a - almost certain': 5,
        'b - likely': 4,
        'c - possible': 3,
        'd - unlikely': 2,
        'e - rare': 1
    }

    consequence_map = {
        '5 - critical': 5,
        '4 - major': 4,
        '3 - moderate': 3,
        '2 - minor': 2,
        '1 - low': 1
    }

    def hv(score):
        if 8 <= score <= 10:
            return 'Extreme'
        elif 6 <= score <= 7:
            return 'High'
        elif score == 5:
            return 'Medium'
        else:
            return 'Low'

    def non_structural(score):
        if score < 7:
            return 'Low Risk'
        elif score < 17:
            return 'Medium Risk'
        else:
            return 'High Risk'

    def structural(score):
        if score < 4:
            return 'Normal'
        elif score < 11:
            return 'Moderate'
        elif score < 20:
            return 'Abnormal'
        else:
            return 'Critical'

    for record in data:
        probability_score = probability_map.get(record.get('probability').lower(), 1)
        consequence_score = consequence_map.get(record.get('consequence').lower(), 1)

        match int(risk_type):
            case 1:
                combined_score = probability_score + consequence_score
                record['risk_rating'] = hv(combined_score)
            case 2:
                combined_score = probability_score * consequence_score
                record['risk_rating'] = non_structural(combined_score)
            case 3:
                combined_score = probability_score * consequence_score
                record['risk_rating'] = structural(combined_score)

    return data


# Column order and display names matching the PDF summary table
EXCEL_COLUMNS = [
    ('date',            'Date'),
    ('area',            'Area'),
    ('unit',            'Unit'),
    ('item',            'Item'),
    ('observations',    'Observations'),
    ('recommendations', 'Recommendations'),
    ('probability',     'Probability'),
    ('consequence',     'Consequence'),
    ('time_ranking',    'Time Ranking'),
    ('risk_rating',     'Risk Rating'),
    ('image_url',       'Image URL'),
]


def generate_excel(data):
    print_memory_usage()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspection Details"

    header_fill = PatternFill(start_color='004d99', end_color='004d99', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    wrap = Alignment(wrap_text=True, vertical='top')

    # Write header row with styling matching PDF
    for col_num, (_, display_name) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=display_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    # Write data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, (db_key, _) in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data.get(db_key))
            cell.alignment = wrap

    # Set column widths
    for col_num in range(1, len(EXCEL_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def filter_by_time(data):
    """
    Sorts a list of dictionaries by 'time_ranking' using a custom priority order.
    """
    ranking_order = [
        'Immediate',
        'Under 1 month',
        'Under 3 months',
        'Under 6 months',
        'Under 12 months',
        'Under 18 months',
        'Over 18 months'
    ]
    ranking_map = {value: index for index, value in enumerate(ranking_order)}
    return sorted(data, key=lambda x: ranking_map.get(x['time_ranking'], float('inf')))


if __name__ == '__main__':
    app.run()
